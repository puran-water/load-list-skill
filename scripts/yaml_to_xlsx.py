#!/usr/bin/env python3
"""
YAML to Excel Converter
Converts load list YAML to formatted Excel workbooks.

Engineering-Grade Outputs (v2.0):
- Load List sheet (all loads with NEC-compliant sizing)
- MCC Schedule sheet (bucket-level detail)
- MCC Panel Summary sheet (panel totals)
- Cable Schedule sheet (for contractor costing)
- Plant Load Summary sheet
- Transformer Schedule sheet
- Energy Summary sheet

Usage:
    python yaml_to_xlsx.py \
        --input electrical/load-list.yaml \
        --output electrical/load-list.xlsx

Author: Load List Skill
Version: 2.0.0
"""

import argparse
import sys
from pathlib import Path
from typing import Optional

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, Border, Side, PatternFill, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# COLUMN DEFINITIONS
# =============================================================================

# Load List columns (v2.0 - with NEC-compliant fields)
LOAD_LIST_COLUMNS = [
    ("equipment_tag", "Tag", 12),
    ("description", "Description", 25),
    ("area", "Area", 8),
    ("rated_kw", "Rated kW", 10),
    ("voltage_v", "Voltage (V)", 11),
    ("frequency_hz", "Freq (Hz)", 10),
    ("motor_poles", "Poles", 7),
    ("efficiency_class", "Eff Class", 10),
    ("efficiency_pct", "Eff (%)", 9),
    ("pf", "PF", 6),
    ("flc_table_a", "FLC (A)", 9),  # NEC table value
    ("fla_nameplate_a", "FLA (A)", 9),  # Nameplate value
    ("lra", "LRA (A)", 9),
    ("service_factor", "SF", 6),
    ("feeder_type", "Feeder Type", 12),
    ("brake_kw", "Brake kW", 10),
    ("absorbed_kw", "Absorbed kW", 12),
    ("duty", "Duty", 10),
    ("running_hours_per_day", "Hrs/Day", 9),
    ("load_factor", "Load Factor", 11),
    ("quantity", "Qty", 6),
    ("diversity_factor", "Diversity", 10),
    ("running_kw", "Running kW", 11),
    ("demand_kw", "Demand kW", 11),
    ("daily_kwh", "kWh/Day", 10),
    ("mcc_panel", "MCC Panel", 11),
]

# MCC Bucket Schedule columns (engineering-grade)
MCC_BUCKET_COLUMNS = [
    ("bucket_id", "Bucket ID", 14),
    ("motor_tag", "Motor Tag", 12),
    ("description", "Description", 25),
    ("motor_kw", "Motor kW", 10),
    ("motor_flc_a", "FLC (A)", 9),
    ("starter_type", "Starter Type", 12),
    ("starter_frame", "Frame", 8),
    ("branch_scpd_type", "SCPD Type", 11),
    ("branch_scpd_rating_a", "SCPD (A)", 10),
    ("overload_setting_a", "OL Set (A)", 10),
    ("overload_class", "OL Class", 9),
    ("bucket_sccr_ka", "SCCR (kA)", 10),
    ("incoming_cable", "In Cable", 12),
    ("outgoing_cable", "Out Cable", 12),
    ("control_voltage", "Ctrl V", 8),
]

# MCC Panel Summary columns
MCC_PANEL_COLUMNS = [
    ("panel_tag", "Panel Tag", 12),
    ("area", "Area", 8),
    ("supply_voltage", "Voltage (V)", 11),
    ("connected_kw", "Connected kW", 13),
    ("running_kw", "Running kW", 12),
    ("demand_kw", "Demand kW", 12),
    ("demand_kva", "Demand kVA", 12),
    ("demand_amps", "Demand (A)", 11),
    ("main_breaker_a", "Main CB (A)", 12),
    ("bus_rating_a", "Bus (A)", 10),
    ("feeder_conductor_min_a", "Fdr Cond (A)", 12),
    ("feeder_ocpd_max_a", "Fdr OCPD (A)", 12),
    ("lineup_sccr_ka", "SCCR (kA)", 10),
    ("bucket_count", "Buckets", 8),
    ("feeder_count_dol", "DOL", 7),
    ("feeder_count_vfd", "VFD", 7),
    ("feeder_count_ss", "Soft Start", 10),
]

# Cable Schedule columns
CABLE_SCHEDULE_COLUMNS = [
    ("cable_tag", "Cable Tag", 14),
    ("from_panel", "From", 12),
    ("to_equipment", "To", 14),
    ("equipment_description", "Description", 25),
    ("motor_kw", "Motor kW", 10),
    ("cable_type", "Cable Type", 12),
    ("cable_construction", "Construction", 25),
    ("cable_size", "Size", 12),
    ("length_m", "Length (m)", 11),
    ("length_assumed", "Assumed", 9),
    ("voltage_drop_pct", "VD (%)", 8),
    ("vd_compliant", "VD OK", 7),
    ("current_a", "Current (A)", 11),
    ("sizing_basis", "Sizing Basis", 25),
]

# Plant Load Summary columns
PLANT_SUMMARY_COLUMNS = [
    ("category", "Category", 25),
    ("connected_kw", "Connected kW", 14),
    ("demand_kw", "Demand kW", 14),
    ("demand_kva", "Demand kVA", 14),
    ("notes", "Notes", 30),
]

# Transformer Schedule columns
TRANSFORMER_COLUMNS = [
    ("xfmr_tag", "Xfmr Tag", 12),
    ("primary_voltage", "Primary V", 11),
    ("secondary_voltage", "Secondary V", 11),
    ("rating_kva", "Rating kVA", 12),
    ("impedance_pct", "Z (%)", 8),
    ("vector_group", "Vector", 10),
    ("served_panels", "Served Panels", 20),
    ("connected_kva", "Connected kVA", 13),
    ("demand_kva", "Demand kVA", 12),
    ("loading_pct", "Loading (%)", 11),
    ("spare_capacity_pct", "Spare (%)", 10),
]

# Key aliases - maps expected column keys to alternate keys from different generators
# Format: expected_key -> [alternate_key1, alternate_key2, ...]
KEY_ALIASES = {
    # Load list keys
    "duty": ["duty_cycle"],
    # MCC Bucket keys
    "motor_kw": ["motor_rated_kw", "rated_kw"],
    "motor_flc_a": ["flc_table_a", "fla"],
    "starter_type": ["unit_type", "feeder_type"],
    "bucket_sccr_ka": ["sccr_ka"],
    "description": ["motor_description", "equipment_description"],
    # Cable Schedule keys
    "vd_compliant": ["voltage_drop_compliant", "compliant_branch"],
    # Transformer keys
    "xfmr_tag": ["transformer_tag"],
    "rating_kva": ["selected_kva"],
    "impedance_pct": ["typical_impedance_pct", "transformer_z_pct"],
    "loading_pct": ["loading_with_growth_pct", "initial_loading_pct"],
    "spare_capacity_pct": ["spare_pct", "margin_pct"],
    "connected_kva": ["connected_load_kva"],
    "demand_kva": ["demand_load_kva"],
    # Panel keys
    "bus_rating_a": ["bus_rating"],
    "feeder_count_dol": ["feeder_counts.DOL"],
    "feeder_count_vfd": ["feeder_counts.VFD"],
    "feeder_count_ss": ["feeder_counts.SOFT_STARTER"],
    "lineup_sccr_ka": ["sccr_ka", "panel_sccr_ka"],
    "feeder_conductor_min_a": ["feeder_conductor_a"],
    "feeder_ocpd_max_a": ["feeder_ocpd_a"],
}


# =============================================================================
# STYLES
# =============================================================================

def create_styles(wb: Workbook) -> dict:
    """Create and register named styles."""
    styles = {}

    # Header style
    header = NamedStyle(name="header")
    header.font = Font(bold=True, color="FFFFFF", size=10)
    header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.border = Border(
        bottom=Side(style="thin", color="000000")
    )
    wb.add_named_style(header)
    styles["header"] = header

    # Data style
    data = NamedStyle(name="data")
    data.font = Font(size=9)
    data.alignment = Alignment(vertical="center")
    wb.add_named_style(data)
    styles["data"] = data

    # Number style
    number = NamedStyle(name="number")
    number.font = Font(size=9)
    number.alignment = Alignment(horizontal="right", vertical="center")
    number.number_format = "#,##0.0"
    wb.add_named_style(number)
    styles["number"] = number

    # Integer style
    integer = NamedStyle(name="integer")
    integer.font = Font(size=9)
    integer.alignment = Alignment(horizontal="right", vertical="center")
    integer.number_format = "#,##0"
    wb.add_named_style(integer)
    styles["integer"] = integer

    # Percent style
    percent = NamedStyle(name="percent")
    percent.font = Font(size=9)
    percent.alignment = Alignment(horizontal="right", vertical="center")
    percent.number_format = "0.0%"
    wb.add_named_style(percent)
    styles["percent"] = percent

    # Title style
    title = NamedStyle(name="title")
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(title)
    styles["title"] = title

    # Subtotal style
    subtotal = NamedStyle(name="subtotal")
    subtotal.font = Font(bold=True, size=9)
    subtotal.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    subtotal.alignment = Alignment(vertical="center")
    wb.add_named_style(subtotal)
    styles["subtotal"] = subtotal

    # Warning style (for non-compliant values)
    warning = NamedStyle(name="warning")
    warning.font = Font(size=9, color="C00000")
    warning.alignment = Alignment(horizontal="right", vertical="center")
    warning.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    wb.add_named_style(warning)
    styles["warning"] = warning

    # Assumed style (for estimated values)
    assumed = NamedStyle(name="assumed")
    assumed.font = Font(size=9, italic=True)
    assumed.alignment = Alignment(vertical="center")
    assumed.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wb.add_named_style(assumed)
    styles["assumed"] = assumed

    return styles


# =============================================================================
# SHEET WRITERS
# =============================================================================

def write_generic_sheet(ws, data: list[dict], columns: list, styles: dict,
                        title: str = None, add_totals: list = None):
    """Generic sheet writer with column definitions."""
    ws.title = title[:31] if title else "Data"  # Excel max 31 chars

    start_row = 1

    # Write headers
    for col_idx, (key, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.style = "header"
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data
    for row_idx, item in enumerate(data, start_row + 1):
        for col_idx, (key, _, _) in enumerate(columns, 1):
            value = item.get(key, "")

            # Try alternate keys if primary key not found
            if value in ("", None) and key in KEY_ALIASES:
                for alt_key in KEY_ALIASES[key]:
                    if "." in alt_key:
                        # Handle nested alternate keys
                        parts = alt_key.split(".")
                        alt_value = item
                        for part in parts:
                            alt_value = alt_value.get(part, "") if isinstance(alt_value, dict) else ""
                        if alt_value not in ("", None):
                            value = alt_value
                            break
                    else:
                        alt_value = item.get(alt_key, "")
                        if alt_value not in ("", None):
                            value = alt_value
                            break

            # Handle nested keys (e.g., "feeder_counts.dol")
            if value in ("", None) and "." in key:
                parts = key.split(".")
                value = item
                for part in parts:
                    value = value.get(part, "") if isinstance(value, dict) else ""

            # Handle special formatting
            if key in ["load_factor", "diversity_factor", "pf", "impedance_pct"]:
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else 0)
                cell.number_format = "0.00"
            elif key in ["efficiency_pct", "loading_pct", "spare_capacity_pct", "voltage_drop_pct"]:
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else 0)
                cell.number_format = "0.0"
            elif key == "length_assumed" and value:
                cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if value else "No")
                cell.style = "assumed" if value else "data"
            elif key == "vd_compliant":
                cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if value else "No")
                cell.style = "data" if value else "warning"
            elif isinstance(value, bool):
                cell = ws.cell(row=row_idx, column=col_idx, value="Yes" if value else "No")
                cell.style = "data"
            elif isinstance(value, (int, float)):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, int):
                    cell.style = "integer"
                else:
                    cell.style = "number"
            elif isinstance(value, list):
                cell = ws.cell(row=row_idx, column=col_idx, value=", ".join(str(v) for v in value))
                cell.style = "data"
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.style = "data"

    # Add totals row if requested
    if add_totals and data:
        total_row = len(data) + start_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").style = "subtotal"
        for col_idx, (key, _, _) in enumerate(columns, 1):
            if key in add_totals:
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(
                    row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{start_row+1}:{col_letter}{total_row-1})"
                )
                cell.style = "subtotal"
                cell.number_format = "#,##0.0"

    # Add autofilter
    if data:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{len(data)+start_row}"

    # Freeze header row
    ws.freeze_panes = f"A{start_row + 1}"


def write_load_list_sheet(ws, loads: list[dict], styles: dict):
    """Write the Load List sheet."""
    write_generic_sheet(
        ws, loads, LOAD_LIST_COLUMNS, styles,
        title="Load List",
        add_totals=["rated_kw", "running_kw", "demand_kw", "daily_kwh"]
    )


def write_mcc_bucket_sheet(ws, buckets: list[dict], styles: dict):
    """Write the MCC Bucket Schedule sheet."""
    write_generic_sheet(
        ws, buckets, MCC_BUCKET_COLUMNS, styles,
        title="MCC Bucket Schedule",
        add_totals=["motor_kw"]
    )


def write_mcc_panel_sheet(ws, panels: list[dict], styles: dict):
    """Write the MCC Panel Summary sheet."""
    # Flatten feeder counts for export
    flat_panels = []
    for panel in panels:
        flat = dict(panel)
        counts = panel.get("feeder_counts", {})
        flat["feeder_count_dol"] = counts.get("dol", 0)
        flat["feeder_count_vfd"] = counts.get("vfd", 0)
        flat["feeder_count_ss"] = counts.get("soft_starter", 0)
        flat_panels.append(flat)

    write_generic_sheet(
        ws, flat_panels, MCC_PANEL_COLUMNS, styles,
        title="MCC Panel Summary",
        add_totals=["connected_kw", "running_kw", "demand_kw", "demand_kva", "bucket_count"]
    )


def write_cable_schedule_sheet(ws, cables: list[dict], styles: dict):
    """Write the Cable Schedule sheet."""
    write_generic_sheet(
        ws, cables, CABLE_SCHEDULE_COLUMNS, styles,
        title="Cable Schedule",
        add_totals=["length_m"]
    )


def write_plant_summary_sheet(ws, summary: dict, styles: dict):
    """Write the Plant Load Summary sheet."""
    ws.title = "Plant Summary"

    # Build summary rows
    rows = []
    s = summary.get("summary", {})

    rows.append({
        "category": "Process Connected Load",
        "connected_kw": s.get("process_connected_kw", 0),
        "demand_kw": s.get("process_demand_kw", 0),
        "demand_kva": s.get("process_demand_kw", 0) / 0.85,
        "notes": "Motor and process equipment"
    })

    rows.append({
        "category": f"Non-Process ({s.get('non_process_allowance_pct', 15)}%)",
        "connected_kw": s.get("non_process_connected_kw", 0),
        "demand_kw": s.get("non_process_demand_kw", 0),
        "demand_kva": s.get("non_process_demand_kw", 0) / 0.90,
        "notes": "HVAC, lighting, controls"
    })

    rows.append({
        "category": "TOTAL (Current)",
        "connected_kw": s.get("total_connected_kw", 0),
        "demand_kw": s.get("total_demand_kw", 0),
        "demand_kva": s.get("total_demand_kva", 0),
        "notes": f"Diversity factor: {s.get('diversity_factor', 0):.2f}"
    })

    fg = summary.get("future_growth", {})
    rows.append({
        "category": f"TOTAL + {fg.get('growth_pct', 20)}% Growth",
        "connected_kw": "",
        "demand_kw": fg.get("future_demand_kw", 0),
        "demand_kva": fg.get("future_demand_kva", 0),
        "notes": "Transformer sizing basis"
    })

    write_generic_sheet(
        ws, rows, PLANT_SUMMARY_COLUMNS, styles,
        title="Plant Summary"
    )


def write_transformer_sheet(ws, transformers: list[dict], styles: dict):
    """Write the Transformer Schedule sheet."""
    write_generic_sheet(
        ws, transformers, TRANSFORMER_COLUMNS, styles,
        title="Transformer Schedule",
        add_totals=["rating_kva", "connected_kva", "demand_kva"]
    )


def write_energy_summary_sheet(ws, summary: dict, styles: dict):
    """Write the Energy Summary sheet."""
    ws.title = "Energy Summary"

    # Summary data
    rows = [
        ("Plant Energy Summary", None, "title"),
        ("", "", ""),
        ("Connected Load (kW)", summary.get("total_connected_kw", 0), "number"),
        ("Running Load (kW)", summary.get("total_running_kw", 0), "number"),
        ("Demand Load (kW)", summary.get("total_demand_kw", 0), "number"),
        ("", "", ""),
        ("Daily Energy (kWh/day)", summary.get("daily_kwh", 0), "number"),
        ("Monthly Energy (MWh/mo)", summary.get("daily_kwh", 0) * 30 / 1000, "number"),
        ("Annual Energy (MWh/yr)", summary.get("daily_kwh", 0) * 365 / 1000, "number"),
        ("", "", ""),
        ("Plant Flow (m³/d)", summary.get("plant_flow_m3_d", 0), "number"),
        ("Specific Energy (kWh/m³)", summary.get("specific_energy_kwh_m3", 0), "number"),
    ]

    for row_idx, (label, value, style) in enumerate(rows, 1):
        if style == "title":
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.style = "title"
            ws.merge_cells(f"A{row_idx}:B{row_idx}")
        else:
            ws.cell(row=row_idx, column=1, value=label).style = "data"
            if value is not None:
                cell = ws.cell(row=row_idx, column=2, value=value)
                if style == "number":
                    cell.number_format = "#,##0.00"

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def write_disclaimer_sheet(ws, disclaimers: list[str], output_tier: int, styles: dict):
    """Write disclaimers and notes sheet."""
    ws.title = "Notes"

    # Title
    ws.cell(row=1, column=1, value="LOAD LIST NOTES AND DISCLAIMERS").style = "title"
    ws.merge_cells("A1:C1")

    # Output tier
    tier_names = {1: "Load Study", 2: "Preliminary Schedule", 3: "Code-Compliant"}
    ws.cell(row=3, column=1, value="Output Tier:")
    ws.cell(row=3, column=2, value=f"Tier {output_tier} - {tier_names.get(output_tier, 'Unknown')}")

    # Disclaimers
    ws.cell(row=5, column=1, value="DISCLAIMERS").style = "header"
    for idx, disclaimer in enumerate(disclaimers, 6):
        ws.cell(row=idx, column=1, value=f"• {disclaimer}").style = "data"

    ws.column_dimensions["A"].width = 80


# =============================================================================
# MAIN CONVERTER
# =============================================================================

def convert_yaml_to_xlsx(
    input_path: Path,
    output_path: Path,
    include_mcc_tabs: bool = True,
    include_cable_schedule: bool = True
):
    """
    Convert load list YAML to Excel workbook.

    Args:
        input_path: Path to load list YAML
        output_path: Path for output Excel file
        include_mcc_tabs: Include per-MCC tabs (default True)
        include_cable_schedule: Include cable schedule sheet (default True)
    """
    # Load YAML
    with open(input_path) as f:
        data = yaml.safe_load(f)

    loads = data.get("loads", [])
    panels = data.get("mcc_panels", [])
    buckets = data.get("mcc_buckets", [])

    # Handle cable_schedule as either list or dict with "cables" key
    cable_data = data.get("cable_schedule", [])
    if isinstance(cable_data, dict):
        cables = cable_data.get("cables", [])
    elif isinstance(cable_data, list):
        cables = cable_data
    else:
        cables = []

    # Handle plant_summary with fallback to plant_load_summary
    plant_summary = data.get("plant_summary", data.get("plant_load_summary", {}))

    transformers = data.get("transformers", [])
    energy_summary = data.get("energy_summary", {})

    # Handle output_tier as either int or object
    output_tier_raw = data.get("output_tier", 1)
    if isinstance(output_tier_raw, dict):
        output_tier = output_tier_raw.get("tier", 1)
        # Get disclaimers from output_tier if present
        disclaimers = output_tier_raw.get("disclaimers", [])
    else:
        output_tier = output_tier_raw if isinstance(output_tier_raw, int) else 1
        disclaimers = []

    # Fallback to top-level disclaimers if not found
    if not disclaimers:
        disclaimers = data.get("disclaimers", [])

    # Create workbook
    wb = Workbook()
    styles = create_styles(wb)

    # 1. Load List sheet (always first)
    write_load_list_sheet(wb.active, loads, styles)

    # 2. MCC Bucket Schedule (if data available)
    if buckets:
        bucket_ws = wb.create_sheet()
        write_mcc_bucket_sheet(bucket_ws, buckets, styles)

    # 3. MCC Panel Summary
    if panels:
        panel_ws = wb.create_sheet()
        write_mcc_panel_sheet(panel_ws, panels, styles)

    # 4. Cable Schedule (if requested and data available)
    if include_cable_schedule and cables:
        cable_ws = wb.create_sheet()
        write_cable_schedule_sheet(cable_ws, cables, styles)

    # 5. Plant Summary (if available)
    if plant_summary:
        summary_ws = wb.create_sheet()
        write_plant_summary_sheet(summary_ws, plant_summary, styles)

    # 6. Transformer Schedule (if available)
    if transformers:
        xfmr_ws = wb.create_sheet()
        write_transformer_sheet(xfmr_ws, transformers, styles)

    # 7. Energy Summary sheet
    if energy_summary:
        energy_ws = wb.create_sheet()
        write_energy_summary_sheet(energy_ws, energy_summary, styles)

    # 8. Notes/Disclaimers sheet
    notes_ws = wb.create_sheet()
    write_disclaimer_sheet(notes_ws, disclaimers, output_tier, styles)

    # 9. Optional: Per-MCC tabs
    if include_mcc_tabs and panels:
        for panel in panels:
            panel_tag = panel.get("panel_tag", "UNKNOWN")
            panel_loads = [l for l in loads if l.get("mcc_panel") == panel_tag]
            if panel_loads:
                # Sanitize sheet name (max 31 chars, no special chars)
                sheet_name = panel_tag[:31].replace("/", "-").replace("\\", "-")
                try:
                    panel_ws = wb.create_sheet(title=sheet_name)
                    write_load_list_sheet(panel_ws, panel_loads, styles)
                except Exception:
                    pass  # Skip if sheet name is invalid

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "loads": len(loads),
        "panels": len(panels),
        "buckets": len(buckets),
        "cables": len(cables)
    }


def main():
    parser = argparse.ArgumentParser(
        description="Convert load list YAML to Excel (v2.0 - Engineering Grade)"
    )
    parser.add_argument(
        "--input", "-i",
        type=Path,
        required=True,
        help="Input YAML file"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        required=True,
        help="Output Excel file"
    )
    parser.add_argument(
        "--no-mcc-tabs",
        action="store_true",
        help="Don't include per-MCC tabs"
    )
    parser.add_argument(
        "--no-cable-schedule",
        action="store_true",
        help="Don't include cable schedule sheet"
    )

    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: Input file not found: {args.input}")
        sys.exit(1)

    print(f"Converting {args.input} to {args.output}...")

    counts = convert_yaml_to_xlsx(
        args.input,
        args.output,
        include_mcc_tabs=not args.no_mcc_tabs,
        include_cable_schedule=not args.no_cable_schedule
    )

    print(f"Done! Exported:")
    print(f"  - {counts['loads']} loads")
    print(f"  - {counts['panels']} MCC panels")
    print(f"  - {counts['buckets']} MCC buckets")
    print(f"  - {counts['cables']} cables")


if __name__ == "__main__":
    main()
